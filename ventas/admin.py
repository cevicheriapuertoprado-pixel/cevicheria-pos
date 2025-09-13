from decimal import Decimal
from django.contrib import admin, messages
from django import forms
from django.urls import path
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from .models import Plato, Pedido, DetallePedido, Caja, Mesa

# ----------------- Formulario para subir Excel -----------------
class UploadExcelForm(forms.Form):
    file = forms.FileField(help_text="Sube un archivo .xlsx. Cada hoja será una categoría.")

# ----------------- Admin personalizado para Plato -----------------
class PlatoAdmin(admin.ModelAdmin):
    list_display = ("nombre", "precio", "categoria", "activo")
    search_fields = ("nombre", "categoria")
    list_filter = ("categoria", "activo")

    # ----------------- URLs personalizadas -----------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel),
                name="ventas_plato_import_excel",
            ),
        ]
        return custom_urls + urls

    # ----------------- Vista para importar Excel -----------------
    def import_excel(self, request):
        if request.method == "POST":
            form = UploadExcelForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data["file"]
                try:
                    wb = load_workbook(filename=file, data_only=True)
                except Exception as e:
                    messages.error(request, f"No se pudo leer el Excel: {e}")
                    return redirect("..")

                creados, actualizados, omitidos, hojas_ignoradas = 0, 0, 0, []

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # Mapeo de encabezados (insensible a mayúsculas)
                    headers = {}
                    for idx, cell in enumerate(ws[1], start=1):
                        if cell.value:
                            headers[str(cell.value).strip().lower()] = idx

                    # Verificar columnas mínimas
                    if "producto" not in headers or "precio" not in headers:
                        hojas_ignoradas.append(sheet_name)
                        continue

                    # Procesar filas
                    for row in ws.iter_rows(min_row=2):
                        nombre = row[headers["producto"] - 1].value
                        precio_val = row[headers["precio"] - 1].value

                        if not nombre:
                            omitidos += 1
                            continue

                        # Parseo robusto de precio
                        try:
                            if isinstance(precio_val, (int, float, Decimal)):
                                precio = Decimal(str(precio_val))
                            else:
                                s = str(precio_val)
                                s = s.replace("S/.", "").replace("S/", "").replace("s/.", "").replace("s/", "")
                                s = s.replace(" ", "").replace(",", ".")
                                filtrado = "".join(ch for ch in s if ch.isdigit() or ch == ".")
                                precio = Decimal(filtrado)
                        except Exception:
                            omitidos += 1
                            continue

                        obj, creado = Plato.objects.update_or_create(
                            nombre=str(nombre).strip(),
                            categoria=sheet_name.strip(),
                            defaults={"precio": precio, "activo": True},
                        )
                        if creado:
                            creados += 1
                        else:
                            actualizados += 1

                # Mensajes finales
                if hojas_ignoradas:
                    messages.warning(request, f"Hojas ignoradas (encabezados faltantes): {', '.join(hojas_ignoradas)}")

                messages.success(
                    request,
                    f"Importación completa ✅ — {creados} creados, {actualizados} actualizados, {omitidos} omitidos."
                )
                return redirect("..")
        else:
            form = UploadExcelForm()

        # Render del formulario (asegúrate de crear este template)
        return render(request, "admin/import_excel.html", {"form": form, "title": "Importar Carta desde Excel"})

# ----------------- Registro de modelos -----------------
admin.site.register(Plato, PlatoAdmin)
admin.site.register(Pedido)
admin.site.register(DetallePedido)
admin.site.register(Caja)
admin.site.register(Mesa)
